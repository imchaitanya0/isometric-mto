"""
GET /api/mto/{job_id}          → poll status + result
GET /api/mto/{job_id}/csv      → download MTO as CSV
GET /api/mto/{job_id}/xlsx     → download MTO as Excel
"""
import csv
import io
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from core.job_store import job_store
from models.mto import JobStatusResponse, MTOResult

router = APIRouter()

CSV_COLUMNS = [
    "item_no", "category", "description", "size_nps", "schedule_rating",
    "material_spec", "end_type", "quantity", "unit", "length_m",
    "confidence", "remarks",
]


def _get_result_or_404(job_id: str) -> MTOResult:
    job = job_store.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' not found.")
    if job["status"] != "completed" or not job["result"]:
        raise HTTPException(status_code=409, detail="Job not completed yet.")
    return job["result"]


@router.get("/mto/{job_id}", response_model=JobStatusResponse)
async def get_mto_status(job_id: str):
    job = job_store.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' not found.")
    return JobStatusResponse(
        job_id=job_id,
        status=job["status"],
        result=job["result"],
        error=job.get("error"),
        filename=job.get("filename"),
    )


@router.get("/mto/{job_id}/csv")
async def download_csv(job_id: str):
    result: MTOResult = _get_result_or_404(job_id)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for item in result.items:
        row = item.model_dump()
        row["end_type"] = row["end_type"] if row["end_type"] else ""
        writer.writerow(row)

    output.seek(0)
    filename = f"mto_{job_id[:8]}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/mto/{job_id}/xlsx")
async def download_xlsx(job_id: str):
    result: MTOResult = _get_result_or_404(job_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "MTO"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, item in enumerate(result.items, start=2):
        data = item.model_dump()
        for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
            val = data.get(col_name)
            if hasattr(val, "value"):  # enum → string
                val = val.value
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"mto_{job_id[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
